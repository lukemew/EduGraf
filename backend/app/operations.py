import pandas as pd
import os
import tempfile
from datetime import datetime
from typing import List
from fastapi import UploadFile
import io

# Importar as funções do sistema original
from app.utils import (
    process_excel_file_real,
    process_excel_file,
    generate_charts_real,
    create_pdf_report,
    debug_dataframe_sections,
    aplicar_formatacao_excel
)

class Operations:
    """Classe simplificada para operações de processamento"""
    
    def __init__(self):
        self.temp_dir = "temp"
        self.upload_dir = "uploads"
        os.makedirs(self.temp_dir, exist_ok=True)
        os.makedirs(self.upload_dir, exist_ok=True)
    
    async def gerar_tabela_do_polo(self, files: List[UploadFile], polo: str = "Geral") -> str:
        """
        Gera tabela consolidada do polo especificado
        
        Args:
            files: Lista de arquivos Excel
            polo: Nome do polo para filtrar
            
        Returns:
            Caminho do arquivo Excel gerado
        """
        try:
            print(f"📊 Processando tabela para polo: {polo}")
            
            # Processar o primeiro arquivo (ou consolidar múltiplos se necessário)
            file = files[0]
            
            # Ler arquivo
            contents = await file.read()
            df = pd.read_excel(io.BytesIO(contents), header=None)
            
            print(f"📄 Arquivo lido: {file.filename} - Shape: {df.shape}")
            
            # Detectar formato
            is_real_format = self._check_if_real_format(df)
            print(f"🔍 Formato detectado: {'Real da Prefeitura' if is_real_format else 'Formato Antigo'}")
            
            if is_real_format:
                # Processar formato real
                processed_data = process_excel_file_real(df, polo)
                output_file = self._save_real_format_table(processed_data, polo)
            else:
                # Processar formato antigo
                # Tentar encontrar cabeçalho
                header_row = self._find_header_row(df)
                if header_row is not None:
                    df = pd.read_excel(io.BytesIO(contents), header=header_row)
                
                processed_data = process_excel_file(df, polo)
                output_file = self._save_old_format_table(processed_data, polo)
            
            print(f"✅ Tabela salva em: {output_file}")
            return output_file
            
        except Exception as e:
            print(f"❌ Erro no processamento da tabela: {e}")
            raise e
    
    async def gerar_graficos(self, files: List[UploadFile], trimestre: int) -> str:
        """
        Gera gráficos e relatório PDF.
        - Se receber 1 arquivo, gera gráficos de período único.
        - Se receber 2 arquivos, gera gráficos comparativos.
        """
        try:
            print(f"📈 Gerando gráficos para o {trimestre}º trimestre")
            from app.utils import generate_charts_real, gerar_grafico_comparativo_periodos, create_pdf_report

            charts_data = {}
            polo_nome = "Geral" # O polo é sempre geral para os gráficos
            total_alunos_p1 = 0
            total_alunos_p2 = None # Inicia como None para o caso de ter só 1 arquivo

            # --- CENÁRIO 1: GRÁFICO DE PERÍODO ÚNICO ---
            if len(files) == 1:
                print("📄 Detectado 1 arquivo. Gerando gráficos de período único.")
                contents = await files[0].read()
                df = pd.read_excel(io.BytesIO(contents), header=None)
                processed_data = process_excel_file_real(df, polo_nome)
                
                # Calcula o total de alunos
                total_alunos_p1 = processed_data['leitura']['total alunos'].sum()
                
                charts_data = generate_charts_real(processed_data, trimestre)

            # --- CENÁRIO 2: GRÁFICO COMPARATIVO ---
            elif len(files) == 2:
                print("📄 Detectados 2 arquivos. Gerando gráficos comparativos por segmento.")
                # Processa Período 1
                contents_p1 = await files[0].read()
                df_p1 = pd.read_excel(io.BytesIO(contents_p1), header=None)
                processed_data_p1 = process_excel_file_real(df_p1, polo_nome)
                
                # Processa Período 2
                contents_p2 = await files[1].read()
                df_p2 = pd.read_excel(io.BytesIO(contents_p2), header=None)
                processed_data_p2 = process_excel_file_real(df_p2, polo_nome)

                # Calcula os totais de alunos de cada período
                total_alunos_p1 = processed_data_p1['leitura']['total alunos'].sum()
                total_alunos_p2 = processed_data_p2['leitura']['total alunos'].sum()

                # Define os grupos de séries
                anos_fund1 = ['1°', '2°', '3°', '4°', '5°', 'EJA SEG I'] # <-- ALTERAÇÃO AQUI
                anos_fund2 = ['6°', '7°', '8°', '9°', 'EJA SEG II']    # <-- ALTERAÇÃO AQUI

                # Não Leitor (NL)
                charts_data['nl_fund1'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='nl', series_selecionadas=anos_fund1, titulo_grafico='Comparativo de Alunos Não Leitores (NL)', subtitulo_grafico='Segmento: Fundamental I (1º ao 5º ano e EJA I)')
                charts_data['nl_fund2'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='nl', series_selecionadas=anos_fund2, titulo_grafico='Comparativo de Alunos Não Leitores (NL)', subtitulo_grafico='Segmento: Fundamental II (6º ao 9º ano e EJA II)')

                # Leitor de Sílabas (LS)
                charts_data['ls_fund1'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='ls', series_selecionadas=anos_fund1, titulo_grafico='Comparativo de Alunos Leitores de sílabas (LS)', subtitulo_grafico='Segmento: Fundamental I (1º ao 5º ano e EJA I)')
                charts_data['ls_fund2'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='ls', series_selecionadas=anos_fund2, titulo_grafico='Comparativo de Alunos Leitores de sílabas (LS)', subtitulo_grafico='Segmento: Fundamental II (6º ao 9º ano e EJA II)')

                # Leitor de Palavras (LP)
                charts_data['lp_fund1'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='lp', series_selecionadas=anos_fund1, titulo_grafico='Comparativo de Alunos Leitores de palavras (LP)', subtitulo_grafico='Segmento: Fundamental I (1º ao 5º ano e EJA I)')
                charts_data['lp_fund2'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='lp', series_selecionadas=anos_fund2, titulo_grafico='Comparativo de Alunos Leitores de palavras (LP)', subtitulo_grafico='Segmento: Fundamental II (6º ao 9º ano e EJA II)')   

                # Leitor de Frases (LF)
                charts_data['lf_fund1'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='lf', series_selecionadas=anos_fund1, titulo_grafico='Comparativo de Alunos Leitores de frases (LF)', subtitulo_grafico='Segmento: Fundamental I (1º ao 5º ano e EJA I)')
                charts_data['lf_fund2'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='lf', series_selecionadas=anos_fund2, titulo_grafico='Comparativo de Alunos Leitores de frases (LF)', subtitulo_grafico='Segmento: Fundamental II (6º ao 9º ano e EJA II)')   

                # Leitor sem Fluência (LSF)
                charts_data['lsf_fund1'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='lsf', series_selecionadas=anos_fund1, titulo_grafico='Comparativo de Alunos Leitores sem fluência (LSF)', subtitulo_grafico='Segmento: Fundamental I (1º ao 5º ano e EJA I)')
                charts_data['lsf_fund2'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='lsf', series_selecionadas=anos_fund2, titulo_grafico='Comparativo de Alunos Leitores sem fluência (LSF)', subtitulo_grafico='Segmento: Fundamental II (6º ao 9º ano e EJA II)')   

                # Leitor com Fluência (LCF)
                charts_data['lcf_fund1'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='lcf', series_selecionadas=anos_fund1, titulo_grafico='Comparativo de Alunos Leitores com fluência (LCF)', subtitulo_grafico='Segmento: Fundamental I (1º ao 5º ano e EJA I)')
                charts_data['lcf_fund2'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='lcf', series_selecionadas=anos_fund2, titulo_grafico='Comparativo de Alunos Leitores com fluência (LCF)', subtitulo_grafico='Segmento: Fundamental II (6º ao 9º ano e EJA II)')   

                # Pré-silábicos (P)
                charts_data['p_fund1'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='p', series_selecionadas=anos_fund1, titulo_grafico='Comparativo de Alunos Pré-Silábicos (P)', subtitulo_grafico='Segmento: Fundamental I (1º ao 5º ano e EJA I)')
                charts_data['p_fund2'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='p', series_selecionadas=anos_fund2, titulo_grafico='Comparativo de Alunos Pré-Silábicos (P)', subtitulo_grafico='Segmento: Fundamental II (6º ao 9º ano e EJA II)')

                # Silábicos (S)
                charts_data['s_fund1'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='s', series_selecionadas=anos_fund1, titulo_grafico='Comparativo de Alunos silábicos (S)', subtitulo_grafico='Segmento: Fundamental I (1º ao 5º ano e EJA I)')
                charts_data['s_fund2'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='s', series_selecionadas=anos_fund2, titulo_grafico='Comparativo de Alunos silábicos (S)', subtitulo_grafico='Segmento: Fundamental II (6º ao 9º ano e EJA II)')
            
                # Silábico Alfabético (S.A)
                charts_data['s.a_fund1'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='s.a.', series_selecionadas=anos_fund1, titulo_grafico='Comparativo de Alunos silábicos alfabéticos (SA)', subtitulo_grafico='Segmento: Fundamental I (1º ao 5º ano e EJA I)')
                charts_data['s.a_fund2'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='s.a.', series_selecionadas=anos_fund2, titulo_grafico='Comparativo de Alunos silábicos alfabéticos (SA)', subtitulo_grafico='Segmento: Fundamental II (6º ao 9º ano e EJA II)')
            
                # Alfabético (A)
                charts_data['a_fund1'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='a', series_selecionadas=anos_fund1, titulo_grafico='Comparativo de Alunos alfabéticos (A)', subtitulo_grafico='Segmento: Fundamental I (1º ao 5º ano e EJA I)')
                charts_data['a_fund2'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='a', series_selecionadas=anos_fund2, titulo_grafico='Comparativo de Alunos alfabéticos (A)', subtitulo_grafico='Segmento: Fundamental II (6º ao 9º ano e EJA II)')
            
                # Ortográficos (O)
                charts_data['o_fund1'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='o', series_selecionadas=anos_fund1, titulo_grafico='Comparativo de Alunos ortográficos (O)', subtitulo_grafico='Segmento: Fundamental I (1º ao 5º ano e EJA I)')
                charts_data['o_fund2'] = gerar_grafico_comparativo_periodos(data_p1=processed_data_p1, data_p2=processed_data_p2, metrica='o', series_selecionadas=anos_fund2, titulo_grafico='Comparativo de Alunos ortográficos (O)', subtitulo_grafico='Segmento: Fundamental II (6º ao 9º ano e EJA II)')
            


            else:
                raise ValueError("Número de arquivos inválido. Envie 1 ou 2 arquivos.")

            # --- Cria o relatório PDF com os gráficos gerados ---
            # CHAMADA CORRIGIDA: Agora passando todos os argumentos
            pdf_path = create_pdf_report(charts_data, trimestre, polo_nome, total_alunos_p1, total_alunos_p2)
            
            print(f"✅ Relatório PDF gerado: {pdf_path}")
            return pdf_path
            
        except Exception as e:
            print(f"❌ Erro na geração de gráficos: {e}")
            import traceback
            traceback.print_exc()
            raise e
    
    async def processar_planilhas_simples(self, files: List[UploadFile]) -> dict:
        """
        Versão simplificada que retorna dados em JSON
        Para compatibilidade com o código original
        
        Args:
            files: Lista de arquivos Excel
            
        Returns:
            Dicionário com dados processados
        """
        try:
            results = []
            
            for file in files:
                print(f"📄 Processando arquivo: {file.filename}")
                
                contents = await file.read()
                df = pd.read_excel(io.BytesIO(contents))
                
                # Processar dados básicos
                file_info = {
                    "filename": file.filename,
                    "rows": len(df),
                    "columns": len(df.columns),
                    "columns_names": df.columns.tolist(),
                    "sample_data": df.head(3).to_dict('records') if len(df) > 0 else []
                }
                
                # Tentar detectar tipo de dados
                if self._has_education_columns(df):
                    file_info["type"] = "dados_educacionais"
                    
                    # Estatísticas básicas se for dados educacionais
                    if 'Nome da escola' in df.columns:
                        file_info["total_escolas"] = df['Nome da escola'].nunique()
                    
                    if 'Niveis de Leitura' in df.columns:
                        file_info["distribuicao_leitura"] = df['Niveis de Leitura'].value_counts().to_dict()
                    
                    if 'Niveis de Escrita' in df.columns:
                        file_info["distribuicao_escrita"] = df['Niveis de Escrita'].value_counts().to_dict()
                else:
                    file_info["type"] = "planilha_generica"
                
                results.append(file_info)
            
            return {
                "total_files": len(files),
                "processed_at": datetime.now().isoformat(),
                "files": results
            }
            
        except Exception as e:
            print(f"❌ Erro no processamento simples: {e}")
            raise e
    
    def _check_if_real_format(self, df: pd.DataFrame) -> bool:
        """Verifica se é formato real da prefeitura"""
        try:
            for idx, row in df.head(50).iterrows():
                for value in row:
                    if isinstance(value, str):
                        value_str = str(value).upper()
                        if "DIAGNÓSTICO DE LEITURA" in value_str or "DIAGNÓSTICO DE ESCRITA" in value_str:
                            return True
                        if "NL" in value_str and "LS" in value_str and "LP" in value_str:
                            return True
            return False
        except:
            return False
    
    def _find_header_row(self, df: pd.DataFrame) -> int:
        """Encontra a linha do cabeçalho no formato antigo"""
        try:
            for idx in range(min(20, len(df))):
                row = df.iloc[idx]
                row_str = ' '.join([str(cell).upper() for cell in row if not pd.isna(cell)])
                
                if "NOME DA ESCOLA" in row_str and "MODALIDADE" in row_str:
                    return idx
            return None
        except:
            return None
    
    def _has_education_columns(self, df: pd.DataFrame) -> bool:
        """Verifica se tem colunas educacionais"""
        education_keywords = [
            'nome da escola', 'modalidade', 'niveis de leitura', 
            'niveis de escrita', 'escola', 'polo', 'diagnóstico'
        ]
        
        columns_str = ' '.join(df.columns.astype(str)).lower()
        return any(keyword in columns_str for keyword in education_keywords)
    
    def _save_real_format_table(self, processed_data: dict, polo: str) -> str:
        """Salva tabela do formato real com formatação profissional"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(
            self.temp_dir, 
            f"tabela_real_{polo.replace(' ', '_')}_{timestamp}.xlsx"
        )
        
        # Criar workbook com openpyxl para formatação avançada
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # Remover planilha padrão
        wb.remove(wb.active)
        
        # Aba de Leitura
        ws_leitura = wb.create_sheet("Dados de Leitura")
        # INSERE OS DADOS A PARTIR DA LINHA 3
        rows_leitura = dataframe_to_rows(processed_data['leitura'], index=False, header=False)
        for r_idx, row in enumerate(rows_leitura, 3):  # O '3' aqui faz começar da linha 3
            for c_idx, value in enumerate(row, 1):
                ws_leitura.cell(row=r_idx, column=c_idx, value=value)
        # AGORA CHAMA A FORMATAÇÃO
        aplicar_formatacao_excel(ws_leitura, processed_data['leitura'], 
                                f" DADOS DE LEITURA - {polo.upper()}", '1abc9c', 'leitura')

        # Aba de Escrita
        ws_escrita = wb.create_sheet("Dados de Escrita")
        # INSERE OS DADOS A PARTIR DA LINHA 3
        rows_escrita = dataframe_to_rows(processed_data['escrita'], index=False, header=False)
        for r_idx, row in enumerate(rows_escrita, 3): # O '3' aqui faz começar da linha 3
            for c_idx, value in enumerate(row, 1):
                ws_escrita.cell(row=r_idx, column=c_idx, value=value)
        # AGORA CHAMA A FORMATAÇÃO
        aplicar_formatacao_excel(ws_escrita, processed_data['escrita'], 
                                f" DADOS DE ESCRITA - {polo.upper()}", '3498db', 'escrita')
        

        # 1. Cálculos Preliminares
        df_leitura = processed_data['leitura']
        df_escrita = processed_data['escrita']

        # Cálculos para Percentuais Gerais
        total_geral_leitura = df_leitura['total alunos'].sum()
        total_geral_escrita = df_escrita['total alunos'].sum()

        # Cálculos para Análise de Progressão (Fund I vs Fund II, incluindo EJA)
        anos_fund1 = ['1°', '2°', '3°', '4°', '5°', 'EJA SEG I'] # <-- ALTERAÇÃO AQUI
        anos_fund2 = ['6°', '7°', '8°', '9°', 'EJA SEG II']    # <-- ALTERAÇÃO AQUI

        # Leitura
        df_fund1_leitura = df_leitura[df_leitura['ano'].isin(anos_fund1)]
        df_fund2_leitura = df_leitura[df_leitura['ano'].isin(anos_fund2)]
        perc_prof_leitura_f1 = (df_fund1_leitura['lcf'].sum() / df_fund1_leitura['total alunos'].sum() * 100) if not df_fund1_leitura.empty and df_fund1_leitura['total alunos'].sum() > 0 else 0
        perc_prof_leitura_f2 = (df_fund2_leitura['lcf'].sum() / df_fund2_leitura['total alunos'].sum() * 100) if not df_fund2_leitura.empty and df_fund2_leitura['total alunos'].sum() > 0 else 0

        # Escrita (Proficientes = Alfabético + Ortográfico)
        df_fund1_escrita = df_escrita[df_escrita['ano'].isin(anos_fund1)]
        df_fund2_escrita = df_escrita[df_escrita['ano'].isin(anos_fund2)]
        soma_prof_escrita_f1 = df_fund1_escrita['a'].sum() + df_fund1_escrita['o'].sum()
        soma_prof_escrita_f2 = df_fund2_escrita['a'].sum() + df_fund2_escrita['o'].sum()
        perc_prof_escrita_f1 = (soma_prof_escrita_f1 / df_fund1_escrita['total alunos'].sum() * 100) if not df_fund1_escrita.empty and df_fund1_escrita['total alunos'].sum() > 0 else 0
        perc_prof_escrita_f2 = (soma_prof_escrita_f2 / df_fund2_escrita['total alunos'].sum() * 100) if not df_fund2_escrita.empty and df_fund2_escrita['total alunos'].sum() > 0 else 0


        # 2. Construção do Dicionário de Dados Refinado
        stats_data = {
            'Métrica': [
                '--- PERCENTUAIS GERAIS ---',
                'Total de Alunos (Leitura)',
                'Total de Alunos (Escrita)',
                'Percentual Geral de Não Leitores (NL)',
                'Percentual Geral de Leitores Fluentes (LCF)',
                'Percentual Geral de Pré-Silábicos (P)',
                'Percentual Geral de Ortográficos (O)',
                '--- ANÁLISE DE PROGRESSÃO ---', # Título atualizado
                '% de Leitores Fluentes (Fund. I + EJA I)',
                '% de Leitores Fluentes (Fund. II + EJA II)',
                '% de Alunos Ortográficos (Fund. I + EJA I)',
                '% de Alunos Ortográficos (Fund. II + EJA II)',
                '--- MÉDIAS POR SÉRIE ---',
                'Média % de Não Leitores por série (NL)',
                'Média % de Leitores Fluentes por série (LCF)',
                'Média % de Pré-Silábicos por série (P)',
                'Média % de Ortográficos por série (O)',
            ],
            'Valor': [
                '', # Linha em branco para separar seções
                f"{total_geral_leitura} (100%)",
                f"{total_geral_escrita} (100%)",
                f"{(df_leitura['nl'].sum() / total_geral_leitura * 100):.1f}%",
                f"{(df_leitura['lcf'].sum() / total_geral_leitura * 100):.1f}%",
                f"{(df_escrita['p'].sum() / total_geral_escrita * 100):.1f}%",
                f"{(df_escrita['o'].sum() / total_geral_escrita * 100):.1f}%",
                '', # Linha em branco
                f"{perc_prof_leitura_f1:.1f}%",
                f"{perc_prof_leitura_f2:.1f}%",
                f"{perc_prof_escrita_f1:.1f}%",
                f"{perc_prof_escrita_f2:.1f}%",
                '', # Linha em branco
                f"{df_leitura['nl_%'].mean():.1f}%",
                f"{df_leitura['lcf_%'].mean():.1f}%",
                f"{df_escrita['p_%'].mean():.1f}%",
                f"{df_escrita['o_%'].mean():.1f}%",
            ]
        }

        # 3. Conversão para DataFrame e escrita na planilha (código que já tínhamos)
        stats_df = pd.DataFrame(stats_data)
        ws_stats = wb.create_sheet("Resumo Estatístico")
        # (O resto do código para escrever o df e chamar a formatação continua o mesmo)
        for c_idx, value in enumerate(stats_df.columns, 1):
            ws_stats.cell(row=2, column=c_idx, value=str(value).title())
        for r_idx, row in enumerate(stats_df.itertuples(), 3):
            ws_stats.cell(row=r_idx, column=1, value=row[1])
            ws_stats.cell(row=r_idx, column=2, value=row[2])
        aplicar_formatacao_excel(ws_stats, stats_df, 
                                f"📈 RESUMO ESTATÍSTICO - {polo.upper()}", 'e74c3c')
        
        # Salvar arquivo
        wb.save(output_file)
        
        return output_file
    
    def _save_old_format_table(self, processed_data: pd.DataFrame, polo: str) -> str:
        """Salva tabela do formato antigo com formatação profissional"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(
            self.temp_dir, 
            f"tabela_antiga_{polo.replace(' ', '_')}_{timestamp}.xlsx"
        )
        
        # Criar workbook com openpyxl para formatação avançada
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # Remover planilha padrão
        wb.remove(wb.active)
        
        # Aba de Dados Consolidados
        ws_dados = wb.create_sheet("Dados Consolidados")
        for r in dataframe_to_rows(processed_data, index=False, header=True):
            ws_dados.append(r)
        aplicar_formatacao_excel(ws_dados, processed_data, 
                               f" DADOS CONSOLIDADOS - {polo.upper()}", '1abc9c', 'leitura')
        
        # Estatísticas básicas
        if not processed_data.empty:
            stats_data = {
                'Métrica': [
                    'Total de Escolas',
                    'Total de Alunos',
                    'Média Alunos/Escola',
                    'Escola com Mais Alunos',
                    'Escola com Menos Alunos'
                ],
                'Valor': [
                    len(processed_data),
                    processed_data['Total Alunos'].sum() if 'Total Alunos' in processed_data.columns else 0,
                    f"{processed_data['Total Alunos'].mean():.1f}" if 'Total Alunos' in processed_data.columns else "N/A",
                    processed_data.loc[processed_data['Total Alunos'].idxmax(), 'Escola'] if 'Total Alunos' in processed_data.columns else "N/A",
                    processed_data.loc[processed_data['Total Alunos'].idxmin(), 'Escola'] if 'Total Alunos' in processed_data.columns else "N/A"
                ]
            }
            
            ws_stats = wb.create_sheet("Resumo Estatístico")

            # Vamos escrever o DataFrame manualmente, nos dando controle total.
            # Escreve o cabeçalho primeiro
            for c_idx, value in enumerate(stats_df.columns, 1):
                ws_stats.cell(row=2, column=c_idx, value=value)

            # Escreve as linhas de dados
            for r_idx, row in enumerate(stats_df.itertuples(), 3):
                # row[1] é a Métrica, row[2] é o Valor
                ws_stats.cell(row=r_idx, column=1, value=row[1])
                ws_stats.cell(row=r_idx, column=2, value=row[2])

            # AGORA, a formatação vai funcionar sobre células que já têm o valor correto.
            aplicar_formatacao_excel(ws_stats, stats_df, 
                                    f"📈 RESUMO ESTATÍSTICO - {polo.upper()}", 'e74c3c')
                    
        # Salvar arquivo
        wb.save(output_file)
        
        return output_file

# A classe Operations está pronta para ser instanciada