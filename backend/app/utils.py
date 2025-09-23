import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from typing import List, Dict, Any
import io
import base64
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, KeepTogether, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import os
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configuração do matplotlib para não usar interface gráfica
plt.switch_backend('Agg')

# Configurar estilo dos gráficos
plt.style.use('default')
sns.set_palette("husl")

def aplicar_formatacao_excel(worksheet, df, titulo, cor_cabecalho='1abc9c', tipo_dados='leitura'):
    """
    Aplica formatação profissional a uma planilha Excel
    
    Args:
        worksheet: Worksheet do openpyxl
        df: DataFrame com os dados
        titulo: Título da planilha
        cor_cabecalho: Cor do cabeçalho (hex)
        tipo_dados: 'leitura' ou 'escrita' para adicionar legendas
    """
    # Cores do projeto EduGraf
    cores = {
        'primaria': '3d626d',
        'secundaria': '165b70', 
        'destaque': '1abc9c',
        'fundo_claro': 'f8f9fa',
        'texto_escuro': '333333',
        'borda': 'dee2e6'
    }
    
    # Estilos de fonte
    fonte_titulo = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    fonte_cabecalho = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    fonte_normal = Font(name='Arial', size=11, color=cores['texto_escuro'])
    fonte_numero = Font(name='Arial', size=11, color=cores['texto_escuro'])
    
    # Estilos de preenchimento
    preenchimento_titulo = PatternFill(start_color=cores['secundaria'], end_color=cores['secundaria'], fill_type='solid')
    preenchimento_cabecalho = PatternFill(start_color=cor_cabecalho, end_color=cor_cabecalho, fill_type='solid')
    preenchimento_linha_par = PatternFill(start_color=cores['fundo_claro'], end_color=cores['fundo_claro'], fill_type='solid')
    
    # Estilos de alinhamento
    alinhamento_centro = Alignment(horizontal='center', vertical='center')
    alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
    alinhamento_direita = Alignment(horizontal='right', vertical='center')
    
    # Estilos de borda
    borda_fina = Border(
        left=Side(style='thin', color=cores['borda']),
        right=Side(style='thin', color=cores['borda']),
        top=Side(style='thin', color=cores['borda']),
        bottom=Side(style='thin', color=cores['borda'])
    )
    
    # Adicionar título
    # Calcular o número de colunas necessárias para o título
    num_colunas = len(df.columns) if len(df.columns) > 0 else 8
    ultima_coluna = chr(65 + num_colunas - 1)  # A, B, C, etc.
    
    worksheet.merge_cells(f'A1:{ultima_coluna}1')
    cell_titulo = worksheet['A1']
    cell_titulo.value = titulo
    cell_titulo.font = fonte_titulo
    cell_titulo.fill = preenchimento_titulo
    cell_titulo.alignment = alinhamento_centro
    
    # Função para formatar nomes das colunas
    def formatar_nome_coluna(nome):
        """Formata o nome da coluna para melhor apresentação"""
        formatacoes = {
            'ano': 'Ano',
            'total alunos': 'Total de Alunos',
            'nl': 'NL',
            'ls': 'LS', 
            'lp': 'LP',
            'lf': 'LF',
            'lsf': 'LSF',
            'lcf': 'LCF',
            'p': 'P',
            's': 'S',
            's.a.': 'S.A.',
            'a': 'A',
            'o': 'O',
            'nl_%': 'NL %',
            'ls_%': 'LS %',
            'lp_%': 'LP %',
            'lf_%': 'LF %',
            'lsf_%': 'LSF %',
            'lcf_%': 'LCF %',
            'p_%': 'P %',
            's_%': 'S %',
            's.a._%': 'S.A. %',
            'a_%': 'A %',
            'o_%': 'O %'
        }
        return formatacoes.get(nome.lower(), nome.title())
    
    # Aplicar formatação aos cabeçalhos
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=2, column=col_num)
        cell.value = formatar_nome_coluna(column_title)
        cell.font = fonte_cabecalho
        cell.fill = preenchimento_cabecalho
        cell.alignment = alinhamento_centro
        cell.border = borda_fina
    
    # Aplicar formatação aos dados
    for row_num in range(3, len(df) + 3):
    # Pega o valor da primeira célula (Métrica) para checar se é um título de seção
        metrica_value = str(worksheet.cell(row=row_num, column=1).value or '')

        # --- SE FOR UM TÍTULO DE SEÇÃO ---
        if "---" in metrica_value:
            # Mescla as células da linha para o título ocupar o espaço da tabela
            worksheet.merge_cells(f'A{row_num}:{ultima_coluna}{row_num}')
            
            cell_titulo_secao = worksheet.cell(row=row_num, column=1)
            # Limpa o texto, removendo os '---' para um visual mais limpo
            cell_titulo_secao.value = metrica_value.replace("---", "").strip()
            cell_titulo_secao.font = fonte_cabecalho # Fonte em negrito e branca
            cell_titulo_secao.fill = preenchimento_titulo # Fundo escuro (mesma cor do título principal)
            cell_titulo_secao.alignment = alinhamento_centro
            cell_titulo_secao.border = borda_fina
            
            # Garante que todas as células na área mesclada tenham a borda
            for col_idx in range(2, num_colunas + 1):
                worksheet.cell(row=row_num, column=col_idx).border = borda_fina

        # --- SE FOR UMA LINHA DE DADOS NORMAL ---
        else:
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                # Aplica formatação padrão (que já existia)
                cell.font = fonte_normal
                cell.border = borda_fina
                if row_num % 2 == 1: # Usar ímpar para a cor não coincidir com a seção
                    cell.fill = preenchimento_linha_par
                
                # Lógica de alinhamento (sugestão: Opção 2 da resposta anterior)
                if col_num == 1:
                    cell.alignment = alinhamento_esquerda
                else:
                    cell.alignment = alinhamento_centro
                
                # Lógica de formatação de números (que já existia)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
    
    # Ajustar largura das colunas com valores mínimos
    larguras_minimas = {
        'A': 8,   # Ano
        'B': 25,  # Total de Alunos
        'C': 10,   # NL, LS, etc.
        'D': 10,   # Percentuais
        'E': 10,
        'F': 10,
        'G': 10,
        'H': 10,
        'I': 10,
        'J': 10,
        'K': 10,
        'L': 10,
        'M': 12,
        'N': 12
    }
    
    for col_idx, column in enumerate(worksheet.columns):
        max_length = 0
        column_letter = None
        
        # Encontrar a primeira célula não mesclada para obter a letra da coluna
        for cell in column:
            try:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            except:
                continue
        
        if column_letter is None:
            continue
            
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Usar largura mínima ou calculada, o que for maior
        largura_minima = larguras_minimas.get(column_letter, 10)
        adjusted_width = max(largura_minima, min(max_length + 3, 25))
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Adicionar filtros
    if len(df) > 0 and len(df.columns) > 0:
        ultima_coluna = chr(65 + len(df.columns) - 1)
        ultima_linha = len(df) + 2
        worksheet.auto_filter.ref = f"A2:{ultima_coluna}{ultima_linha}"
    
    # Adicionar legendas explicativas
    linha_legenda = len(df) + 4  # 2 linhas após os dados
    
    # Título da legenda
    cell_legenda_titulo = worksheet.cell(row=linha_legenda, column=1)
    cell_legenda_titulo.value = " LEGENDA - SIGNIFICADO DAS ABREVIAÇÕES"
    cell_legenda_titulo.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    cell_legenda_titulo.fill = PatternFill(start_color=cores['secundaria'], end_color=cores['secundaria'], fill_type='solid')
    cell_legenda_titulo.alignment = alinhamento_centro
    
    # Mesclar células do título da legenda
    ultima_coluna_legenda = chr(65 + min(len(df.columns) - 1, 6))  # Máximo 7 colunas
    worksheet.merge_cells(f'A{linha_legenda}:{ultima_coluna_legenda}{linha_legenda}')
    
    if tipo_dados == 'leitura':
        # Legendas para leitura
        legendas_leitura = [
            ("NL", "Não Leitor - Aluno que não consegue ler"),
            ("LS", "Leitor Silábico - Lê sílaba por sílaba"),
            ("LP", "Leitor Palavra - Lê palavra por palavra"),
            ("LF", "Leitor Frase - Lê frase por frase"),
            ("LSF", "Leitor Silábico com Fluência - Lê sílabas com fluência"),
            ("LCF", "Leitor com Fluência - Lê com fluência total")
        ]
        
        for idx, (abrev, descricao) in enumerate(legendas_leitura):
            linha_atual = linha_legenda + 1 + idx
            
            # Abreviação
            cell_abrev = worksheet.cell(row=linha_atual, column=1)
            cell_abrev.value = abrev
            cell_abrev.font = Font(name='Arial', size=11, bold=True, color=cores['texto_escuro'])
            cell_abrev.fill = PatternFill(start_color='f0f8ff', end_color='f0f8ff', fill_type='solid')
            cell_abrev.alignment = alinhamento_centro
            cell_abrev.border = borda_fina
            
            # Descrição
            cell_desc = worksheet.cell(row=linha_atual, column=2)
            cell_desc.value = descricao
            cell_desc.font = Font(name='Arial', size=10, color=cores['texto_escuro'])
            cell_desc.fill = PatternFill(start_color='f0f8ff', end_color='f0f8ff', fill_type='solid')
            cell_desc.alignment = alinhamento_esquerda
            cell_desc.border = borda_fina
            
            # Mesclar células da descrição
            worksheet.merge_cells(f'B{linha_atual}:{ultima_coluna_legenda}{linha_atual}')
    
    elif tipo_dados == 'escrita':
        # Legendas para escrita
        legendas_escrita = [
            ("P", "Pré-Silábico - Não relaciona som e letra"),
            ("S", "Silábico - Escreve uma letra por sílaba"),
            ("S.A.", "Silábico Alfabético - Mistura sílaba e letra"),
            ("A", "Alfabético - Escreve todas as letras"),
            ("O", "Ortográfico - Escreve corretamente")
        ]
        
        for idx, (abrev, descricao) in enumerate(legendas_escrita):
            linha_atual = linha_legenda + 1 + idx
            
            # Abreviação
            cell_abrev = worksheet.cell(row=linha_atual, column=1)
            cell_abrev.value = abrev
            cell_abrev.font = Font(name='Arial', size=11, bold=True, color=cores['texto_escuro'])
            cell_abrev.fill = PatternFill(start_color='f0f8ff', end_color='f0f8ff', fill_type='solid')
            cell_abrev.alignment = alinhamento_centro
            cell_abrev.border = borda_fina
            
            # Descrição
            cell_desc = worksheet.cell(row=linha_atual, column=2)
            cell_desc.value = descricao
            cell_desc.font = Font(name='Arial', size=10, color=cores['texto_escuro'])
            cell_desc.fill = PatternFill(start_color='f0f8ff', end_color='f0f8ff', fill_type='solid')
            cell_desc.alignment = alinhamento_esquerda
            cell_desc.border = borda_fina
            
            # Mesclar células da descrição
            worksheet.merge_cells(f'B{linha_atual}:{ultima_coluna_legenda}{linha_atual}')
    
    # Congelar painéis (cabeçalho sempre visível)
    worksheet.freeze_panes = 'A3'
    print("Ajustando dimensões personalizadas...")
    if "RESUMO ESTATÍSTICO" in titulo.upper():
        print("Aplicando dimensões personalizadas para a aba de Resumo Estatístico...")

        # LARGURA DAS COLUNAS para a aba de estatísticas
        worksheet.column_dimensions['A'].width = 45  # Coluna 'Métrica'
        worksheet.column_dimensions['B'].width = 25  # Coluna 'Valor'

        # ALTURA DAS LINHAS para a aba de estatísticas
        worksheet.row_dimensions[1].height = 40  # Linha do Título Principal
        worksheet.row_dimensions[2].height = 30  # Linha do Cabeçalho ('Métrica', 'Valor')
    
    else:
        # Dimensões para as outras abas (Leitura e Escrita)
        print(f"Aplicando dimensões padrão para a aba '{titulo}'...")
        worksheet.column_dimensions['A'].width = 15 # Coluna 'Ano'
        worksheet.column_dimensions['B'].width = 20 # Coluna 'Total de Alunos'


    print(f"--- FIM DO DIAGNÓSTICO PARA A ABA: '{titulo}' ---\n")
    
    return worksheet

def criar_dataframe(arquivo_em_bytes, linhas_a_serem_puladas, colunas_a_serem_usadas):
    """
    Cria DataFrame a partir de bytes do arquivo Excel
    
    Args:
        arquivo_em_bytes: Bytes do arquivo Excel
        linhas_a_serem_puladas: Lista de linhas para pular
        colunas_a_serem_usadas: Lista de colunas para usar
    
    Returns:
        DataFrame processado
    """
    df = pd.read_excel(BytesIO(arquivo_em_bytes), 
                       skiprows=linhas_a_serem_puladas,
                       usecols=colunas_a_serem_usadas)
    return df

def process_excel_file_real(df: pd.DataFrame, polo: str) -> Dict[str, Any]:
    """
    Processa planilha Excel real da prefeitura e consolida dados
    CORRIGIDO: Agora lê todas as linhas disponíveis, não apenas 5
    
    Args:
        df: DataFrame com os dados da planilha
        polo: Nome do polo para filtrar
    
    Returns:
        Dicionário com dados processados
    """
    try:
        print(f"🔍 DEBUG: Iniciando processamento do formato real")
        print(f"🔍 DEBUG: DataFrame original shape: {df.shape}")
        
        # Configurações para leitura (suas configurações originais)
        linhas_para_pular_leitura = [0, 1, 2, 3, 4, 5, 12, 17, 20, 21, 22, *range(23, 45)]
        linhas_para_pular_escrita = [*range(0, 27), 33, 38, 41, 42, 43, 44]
        cols_leitura = list(range(0, 14))
        cols_escrita = list(range(0, 12))
        
        columns_leitura = ["ano", "total alunos", "nl", "nl%", "ls", "ls%", "lp", "lp%", "lf", "lf%", "lsf", "lsf%", "lcf", "lcf%"]
        columns_escrita = ["ano", "total alunos", "p", "p%", "s", "s%", "s.a.", "s.a%", "a", "a%", "o", "o%"]
        
        print(f"🔍 DEBUG: Linhas para pular leitura: {linhas_para_pular_leitura}")
        print(f"🔍 DEBUG: Linhas para pular escrita: {linhas_para_pular_escrita}")
        
        # SEÇÃO DE LEITURA - Identificar todas as linhas válidas
        print(f"🔍 DEBUG: === PROCESSANDO SEÇÃO DE LEITURA ===")
        
        # Encontrar a linha inicial da seção de leitura (procurar por indicadores)
        inicio_leitura = None
        for idx in range(min(50, len(df))):
            if idx >= len(df):
                break
            row = df.iloc[idx]
            for value in row:
                if pd.notna(value):
                    value_str = str(value).upper()
                    if "DIAGNÓSTICO DE LEITURA" in value_str or "LEITURA" in value_str:
                        inicio_leitura = idx
                        print(f"🔍 DEBUG: Seção de leitura encontrada na linha {idx}")
                        break
            if inicio_leitura is not None:
                break
        
        # Se não encontrar, usar posição padrão
        if inicio_leitura is None:
            inicio_leitura = 6
            print(f"🔍 DEBUG: Usando posição padrão para leitura: {inicio_leitura}")
        
        # Ler dados de leitura - encontrar todas as linhas com dados válidos
        df_leitura_raw = []
        max_linha_leitura = min(inicio_leitura + 50, len(df))  # Buscar nas próximas 50 linhas
        
        for idx in range(inicio_leitura, max_linha_leitura):
            if idx in linhas_para_pular_leitura:
                continue
                
            if idx >= len(df):
                break
                
            row = df.iloc[idx, :len(cols_leitura)]  # Pegar apenas as colunas necessárias
            
            # Verificar se a linha tem dados válidos (pelo menos o primeiro campo não vazio)
            if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip() != '':
                # Verificar se parece ser uma linha de dados (tem números ou anos)
                primeiro_valor = str(row.iloc[0]).strip()
                pv_upper = primeiro_valor.upper()

                # A nova lógica é: a linha é válida se começar com 'EJA', ou se o primeiro caractere for um número,
                # ou se for a palavra exata 'ANO' ou 'SÉRIE' (para pegar o cabeçalho).
                if (pv_upper.startswith('EJA') or
                    (len(primeiro_valor) > 0 and primeiro_valor[0].isdigit()) or
                    pv_upper == 'ANO' or
                    pv_upper == 'SÉRIE'):
                    df_leitura_raw.append(row.tolist())
                    print(f"🔍 DEBUG: Linha de leitura válida {idx}: {primeiro_valor}")
        
        print(f"🔍 DEBUG: Total de linhas de leitura encontradas: {len(df_leitura_raw)}")
        
        # Criar DataFrame de leitura
        if df_leitura_raw:
            df_leitura = pd.DataFrame(df_leitura_raw, columns=columns_leitura[:len(df_leitura_raw[0])])
            # Completar colunas faltantes se necessário
            for col in columns_leitura:
                if col not in df_leitura.columns:
                    df_leitura[col] = 0
            df_leitura = df_leitura[columns_leitura]  # Reordenar colunas
        else:
            # Fallback para método original se não encontrar dados
            print("🔍 DEBUG: Usando método fallback para leitura")
            df_leitura = df.iloc[inicio_leitura:inicio_leitura+15, 0:14].copy()
            df_leitura.columns = columns_leitura
        
        print(f"🔍 DEBUG: DataFrame de leitura shape: {df_leitura.shape}")
        print(f"🔍 DEBUG: Primeiras linhas de leitura:\n{df_leitura.head()}")
        
        # SEÇÃO DE ESCRITA - Identificar todas as linhas válidas
        print(f"🔍 DEBUG: === PROCESSANDO SEÇÃO DE ESCRITA ===")
        
        # Encontrar a linha inicial da seção de escrita
        inicio_escrita = None
        for idx in range(min(50, len(df))):
            if idx >= len(df):
                break
            row = df.iloc[idx]
            for value in row:
                if pd.notna(value):
                    value_str = str(value).upper()
                    if "DIAGNÓSTICO DE ESCRITA" in value_str or ("ESCRITA" in value_str and "LEITURA" not in value_str):
                        inicio_escrita = idx
                        print(f"🔍 DEBUG: Seção de escrita encontrada na linha {idx}")
                        break
            if inicio_escrita is not None:
                break
        
        # Se não encontrar, usar posição padrão
        if inicio_escrita is None:
            inicio_escrita = 28
            print(f"🔍 DEBUG: Usando posição padrão para escrita: {inicio_escrita}")
        
        # Ler dados de escrita - encontrar todas as linhas com dados válidos
        df_escrita_raw = []
        max_linha_escrita = min(inicio_escrita + 50, len(df))  # Buscar nas próximas 50 linhas
        
        for idx in range(inicio_escrita, max_linha_escrita):
            if idx in linhas_para_pular_escrita:
                continue
                
            if idx >= len(df):
                break
                
            row = df.iloc[idx, :len(cols_escrita)]  # Pegar apenas as colunas necessárias
            
            # Verificar se a linha tem dados válidos
            if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip() != '':
                primeiro_valor = str(row.iloc[0]).strip()
                pv_upper = primeiro_valor.upper()

                # A nova lógica é: a linha é válida se começar com 'EJA', ou se o primeiro caractere for um número,
                # ou se for a palavra exata 'ANO' ou 'SÉRIE' (para pegar o cabeçalho).
                if (pv_upper.startswith('EJA') or
                    (len(primeiro_valor) > 0 and primeiro_valor[0].isdigit()) or
                    pv_upper == 'ANO' or
                    pv_upper == 'SÉRIE'):

                    df_escrita_raw.append(row.tolist())
                    print(f"🔍 DEBUG: Linha de escrita válida {idx}: {primeiro_valor}")
        
        print(f"🔍 DEBUG: Total de linhas de escrita encontradas: {len(df_escrita_raw)}")
        
        # Criar DataFrame de escrita
        if df_escrita_raw:
            df_escrita = pd.DataFrame(df_escrita_raw, columns=columns_escrita[:len(df_escrita_raw[0])])
            # Completar colunas faltantes se necessário
            for col in columns_escrita:
                if col not in df_escrita.columns:
                    df_escrita[col] = 0
            df_escrita = df_escrita[columns_escrita]  # Reordenar colunas
        else:
            # Fallback para método original se não encontrar dados
            print("🔍 DEBUG: Usando método fallback para escrita")
            df_escrita = df.iloc[inicio_escrita:inicio_escrita+15, 0:12].copy()
            df_escrita.columns = columns_escrita
        
        print(f"🔍 DEBUG: DataFrame de escrita shape: {df_escrita.shape}")
        print(f"🔍 DEBUG: Primeiras linhas de escrita:\n{df_escrita.head()}")
        
        # PROCESSAMENTO DOS DADOS
        print(f"🔍 DEBUG: === PROCESSANDO DADOS NUMÉRICOS ===")
        
        # Converter colunas numéricas para leitura
        numeric_cols_leitura = ['total alunos', 'nl', 'ls', 'lp', 'lf', 'lsf', 'lcf']
        numeric_cols_escrita = ['total alunos', 'p', 's', 's.a.', 'a', 'o']
        
        for col in numeric_cols_leitura:
            if col in df_leitura.columns:
                df_leitura[col] = pd.to_numeric(df_leitura[col], errors='coerce')
        
        for col in numeric_cols_escrita:
            if col in df_escrita.columns:
                df_escrita[col] = pd.to_numeric(df_escrita[col], errors='coerce')
        
        # Preencher NaN com 0
        df_leitura = df_leitura.fillna(0).infer_objects(copy=False)
        df_escrita = df_escrita.fillna(0).infer_objects(copy=False)
        
        # Remover linhas completamente vazias (onde todas as colunas numéricas são 0)
        df_leitura = df_leitura[df_leitura[numeric_cols_leitura].sum(axis=1) > 0]
        df_escrita = df_escrita[df_escrita[numeric_cols_escrita].sum(axis=1) > 0]
        
        print(f"🔍 DEBUG: Após limpeza - Leitura: {len(df_leitura)} linhas, Escrita: {len(df_escrita)} linhas")
        
        # Processar dados de leitura - agrupar se necessário
        if 'ano' in df_leitura.columns:
            df_leitura_soma = df_leitura.groupby('ano').agg({
                col: 'sum' for col in numeric_cols_leitura
            }).reset_index()
        else:
            df_leitura_soma = df_leitura.copy()
        
        # Recalcular percentuais de leitura
        if len(df_leitura_soma) > 0:
            total_col = df_leitura_soma['total alunos'] 
            df_leitura_soma['nl_%'] = (df_leitura_soma['nl'] / total_col * 100).fillna(0).round(2)
            df_leitura_soma['ls_%'] = (df_leitura_soma['ls'] / total_col * 100).fillna(0).round(2)
            df_leitura_soma['lp_%'] = (df_leitura_soma['lp'] / total_col * 100).fillna(0).round(2)
            df_leitura_soma['lf_%'] = (df_leitura_soma['lf'] / total_col * 100).fillna(0).round(2)
            df_leitura_soma['lsf_%'] = (df_leitura_soma['lsf'] / total_col * 100).fillna(0).round(2)
            df_leitura_soma['lcf_%'] = (df_leitura_soma['lcf'] / total_col * 100).fillna(0).round(2)
        
        # Processar dados de escrita - agrupar se necessário
        if 'ano' in df_escrita.columns:
            df_escrita_soma = df_escrita.groupby('ano').agg({
                col: 'sum' for col in numeric_cols_escrita
            }).reset_index()
        else:
            df_escrita_soma = df_escrita.copy()
        
        # Recalcular percentuais de escrita
        if len(df_escrita_soma) > 0:
            total_col = df_escrita_soma['total alunos']
            df_escrita_soma['p_%'] = (df_escrita_soma['p'] / total_col * 100).fillna(0).round(2)
            df_escrita_soma['s_%'] = (df_escrita_soma['s'] / total_col * 100).fillna(0).round(2)
            df_escrita_soma['s.a._%'] = (df_escrita_soma['s.a.'] / total_col * 100).fillna(0).round(2)
            df_escrita_soma['a_%'] = (df_escrita_soma['a'] / total_col * 100).fillna(0).round(2)
            df_escrita_soma['o_%'] = (df_escrita_soma['o'] / total_col * 100).fillna(0).round(2)
        
        print(f"🔍 DEBUG: === RESULTADO FINAL ===")
        print(f"🔍 DEBUG: Linhas finais de leitura: {len(df_leitura_soma)}")
        print(f"🔍 DEBUG: Linhas finais de escrita: {len(df_escrita_soma)}")
        print(f"🔍 DEBUG: Dados de leitura:\n{df_leitura_soma}")
        print(f"🔍 DEBUG: Dados de escrita:\n{df_escrita_soma}")
        
        return {
            'leitura': df_leitura_soma,
            'escrita': df_escrita_soma,
            'polo': polo
        }
        
    except Exception as e:
        print(f"🔍 DEBUG: Erro no processamento da planilha real: {str(e)}")
        import traceback
        traceback.print_exc()
        raise Exception(f"Erro no processamento da planilha real: {str(e)}")

# Função auxiliar para debugging melhorado
def debug_dataframe_sections(df: pd.DataFrame):
    """Debug detalhado das seções do DataFrame"""
    print(f"🔍 DEBUG: === ANÁLISE DETALHADA DO DATAFRAME ===")
    print(f"🔍 DEBUG: Shape total: {df.shape}")
    
    # Procurar seções
    secoes_encontradas = []
    
    for idx in range(len(df)):
        if idx >= len(df):
            break
            
        row = df.iloc[idx]
        for col_idx, value in enumerate(row):
            if pd.notna(value):
                value_str = str(value).upper()
                
                if "LEITURA" in value_str:
                    secoes_encontradas.append(f"LEITURA na linha {idx}, coluna {col_idx}: '{value}'")
                elif "ESCRITA" in value_str:
                    secoes_encontradas.append(f"ESCRITA na linha {idx}, coluna {col_idx}: '{value}'")
                elif any(ano in value_str for ano in ["1° ANO", "2° ANO", "3° ANO", "4° ANO", "5° ANO", "6° ANO", "7° ANO", "8° ANO", "9° ANO"]):
                    secoes_encontradas.append(f"ANO na linha {idx}, coluna {col_idx}: '{value}'")
    
    print(f"🔍 DEBUG: Seções encontradas:")
    for secao in secoes_encontradas[:20]:  # Mostrar apenas as primeiras 20
        print(f"🔍 DEBUG:   {secao}")
    
    # Mostrar estatísticas por linha
    print(f"🔍 DEBUG: === ESTATÍSTICAS POR LINHA (primeiras 50) ===")
    for idx in range(min(50, len(df))):
        if idx >= len(df):
            break
            
        row = df.iloc[idx]
        valores_nao_vazios = [str(v)[:20] for v in row if pd.notna(v) and str(v).strip() != '']
        
        if valores_nao_vazios:
            print(f"🔍 DEBUG: L{idx:2d}: {' | '.join(valores_nao_vazios[:3])}{'...' if len(valores_nao_vazios) > 3 else ''}")


def process_excel_file(df: pd.DataFrame, polo: str) -> pd.DataFrame:
    """
    Processa planilha Excel e consolida dados por escola e polo (formato antigo)
    
    Args:
        df: DataFrame com os dados da planilha
        polo: Nome do polo para filtrar
    
    Returns:
        DataFrame processado com dados consolidados
    """
    try:
        # Limpar dados
        df = df.dropna()
        
        # Validar colunas
        required_columns = ['Nome da escola', 'Modalidade', 'Niveis de Leitura', 'Niveis de Escrita']
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"Coluna obrigatória '{col}' não encontrada")
        
        # Filtrar por polo (se necessário)
        if polo and polo != "Geral":
            df = df[df['Nome da escola'].str.contains(polo, case=False, na=False)]
        
        # Calcular estatísticas por escola
        consolidated_data = []
        
        for escola in df['Nome da escola'].unique():
            escola_data = df[df['Nome da escola'] == escola]
            
            # Calcular médias e totais
            total_alunos = len(escola_data)
            
            # Níveis de Leitura
            leitura_stats = escola_data['Niveis de Leitura'].value_counts()
            leitura_baixo = leitura_stats.get('Baixo', 0)
            leitura_medio = leitura_stats.get('Médio', 0)
            leitura_alto = leitura_stats.get('Alto', 0)
            
            # Níveis de Escrita
            escrita_stats = escola_data['Niveis de Escrita'].value_counts()
            escrita_baixo = escrita_stats.get('Baixo', 0)
            escrita_medio = escrita_stats.get('Médio', 0)
            escrita_alto = escrita_stats.get('Alto', 0)
            
            # Calcular percentuais
            leitura_baixo_pct = (leitura_baixo / total_alunos * 100) if total_alunos > 0 else 0
            leitura_medio_pct = (leitura_medio / total_alunos * 100) if total_alunos > 0 else 0
            leitura_alto_pct = (leitura_alto / total_alunos * 100) if total_alunos > 0 else 0
            
            escrita_baixo_pct = (escrita_baixo / total_alunos * 100) if total_alunos > 0 else 0
            escrita_medio_pct = (escrita_medio / total_alunos * 100) if total_alunos > 0 else 0
            escrita_alto_pct = (escrita_alto / total_alunos * 100) if total_alunos > 0 else 0
            
            # Modalidade mais comum
            modalidade_mais_comum = escola_data['Modalidade'].mode().iloc[0] if not escola_data['Modalidade'].mode().empty else "N/A"
            
            consolidated_data.append({
                'Nº': len(consolidated_data) + 1,
                'Escola': escola,
                'Modalidade': modalidade_mais_comum,
                'Total Alunos': total_alunos,
                'Nível de Leitura - Baixo': f"{leitura_baixo} ({leitura_baixo_pct:.1f}%)",
                'Nível de Leitura - Médio': f"{leitura_medio} ({leitura_medio_pct:.1f}%)",
                'Nível de Leitura - Alto': f"{leitura_alto} ({leitura_alto_pct:.1f}%)",
                'Nível de Escrita - Baixo': f"{escrita_baixo} ({escrita_baixo_pct:.1f}%)",
                'Nível de Escrita - Médio': f"{escrita_medio} ({escrita_medio_pct:.1f}%)",
                'Nível de Escrita - Alto': f"{escrita_alto} ({escrita_alto_pct:.1f}%)"
            })
        
        return pd.DataFrame(consolidated_data)
        
    except Exception as e:
        raise Exception(f"Erro no processamento da planilha: {str(e)}")

def gerar_grafico_de_periodo_unico(data_p1: Dict[str, Any], series_selecionadas: List[str], metrica: str, titulo_grafico: str, subtitulo_grafico: str) -> Dict[str, str]:
    """
    Gera um gráfico de barras segmentado para um único período, com o novo estilo visual.
    """
    try:
        leitura_metrics = ['nl', 'ls', 'lp', 'lf', 'lsf', 'lcf']
        escrita_metrics = ['p', 's', 's.a.', 'a', 'o']

        # Seleciona o DataFrame correto (leitura ou escrita)
        if metrica in leitura_metrics:
            df = data_p1['leitura']
        elif metrica in escrita_metrics:
            df = data_p1['escrita']
        else:
            raise ValueError(f"Métrica '{metrica}' desconhecida.")

        # Filtra pelas séries desejadas
        df_final = df[df['ano'].isin(series_selecionadas)].reset_index(drop=True)

        if df_final.empty:
            print(f"⚠️ Aviso: Nenhum dado encontrado para as séries {series_selecionadas} na métrica '{metrica}'.")
            return {}

        # Prepara os dados para plotagem
        categorias = [f"{row['ano']}\nTotal: {int(row['total alunos'])}" for _, row in df_final.iterrows()]
        valores_periodo = df_final[metrica]

        # --- CÓDIGO DE ESTILIZAÇÃO (baseado no gráfico comparativo) ---
        fig, ax = plt.subplots(figsize=(12, 7))
        fig.patch.set_facecolor('white')
        x = np.arange(len(categorias))
        width = 0.45 # Barras um pouco mais largas

        # Desenha apenas as barras do 1º período (verdes)
        bars1 = ax.bar(x, valores_periodo, width, color='#02984c', edgecolor='white', linewidth=1)

        def autolabel(bars):
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax.text(bar.get_x() + bar.get_width()/2.0, height, f'{int(height)}',
                            ha='center', va='bottom', fontsize=10, fontweight='bold')
        
        autolabel(bars1)

        # Configurações visuais
        ax.set_title('')
        ax.set_xlabel('')
        ax.set_ylabel('')
        ax.set_xticks(x)
        ax.set_xticklabels(categorias, rotation=0, ha='center', fontsize=10)
        
        max_value = max(valores_periodo, default=0)
        ax.set_ylim(0, max_value * 1.25 if max_value > 0 else 10)

        ax.grid(True, alpha=0.5, linestyle='-', axis='y', color='lightgray')
        ax.set_axisbelow(True)
        
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.spines['bottom'].set_color('lightgray')

        plt.tight_layout()

        # Salva o arquivo
        os.makedirs("temp", exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        chart_path = f"temp/grafico_unico_{metrica}_{''.join(series_selecionadas)}_{timestamp}.png"
        fig.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)

        print(f"✅ Gráfico de período único '{titulo_grafico}' salvo em: {chart_path}")
        
        return { 'path': chart_path, 'title': titulo_grafico, 'description': subtitulo_grafico }
    except Exception as e:
        print(f"❌ Erro ao gerar gráfico de período único para métrica '{metrica}': {e}")
        import traceback
        traceback.print_exc()
        return {}


# Função para GRÁFICOS COMPARATIVOS DE 2 PERÍODOS
def gerar_grafico_comparativo_periodos(data_p1: Dict[str, Any], data_p2: Dict[str, Any], metrica: str, series_selecionadas: List[str], titulo_grafico: str, subtitulo_grafico: str) -> Dict[str, str]:
    """
    Gera um gráfico de barras comparando NÚMEROS ABSOLUTOS de uma métrica 
    específica entre dois períodos para um grupo de séries.
    """
    try:
        leitura_metrics = ['nl', 'ls', 'lp', 'lf', 'lsf', 'lcf']
        escrita_metrics = ['p', 's', 's.a.', 'a', 'o']

        if metrica in leitura_metrics:
            df1 = data_p1['leitura']
            df2 = data_p2['leitura']
        elif metrica in escrita_metrics:
            df1 = data_p1['escrita']
            df2 = data_p2['escrita']
        else:
            raise ValueError(f"Métrica '{metrica}' desconhecida.")

        df_merged = pd.merge(df1[['ano', 'total alunos', metrica]], df2[['ano', 'total alunos', metrica]], on='ano', suffixes=('_p1', '_p2'), how='inner')
        df_final = df_merged[df_merged['ano'].isin(series_selecionadas)].reset_index(drop=True)

        if df_final.empty:
            print(f"⚠️ Aviso: Nenhum dado encontrado para as séries {series_selecionadas} na métrica '{metrica}'.")
            return {}

        categorias = [f"{row['ano']}\nTotal: {int(row['total alunos_p1'])} - {int(row['total alunos_p2'])}" for _, row in df_final.iterrows()]
        valores_1periodo = df_final[f'{metrica}_p1']
        valores_2periodo = df_final[f'{metrica}_p2']

        fig, ax = plt.subplots(figsize=(12, 7))
        fig.patch.set_facecolor('white')
        x = np.arange(len(categorias))
        width = 0.35
        bars1 = ax.bar(x - width/2, valores_1periodo, width, color='#02984c', label='1º Período', edgecolor='white', linewidth=1)
        bars2 = ax.bar(x + width/2, valores_2periodo, width, color='#fdd835', label='2º Período', edgecolor='white', linewidth=1)

        def autolabel(bars):
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax.text(bar.get_x() + bar.get_width()/2.0, height, f'{int(height)}', ha='center', va='bottom', fontsize=10, fontweight='bold')
        
        autolabel(bars1)
        autolabel(bars2)

        ax.set_title('')
        ax.set_xlabel('')
        ax.set_ylabel('')
        ax.set_xticks(x)
        ax.set_xticklabels(categorias, rotation=0, ha='center', fontsize=10)
        
        max_value = max(max(valores_1periodo, default=0), max(valores_2periodo, default=0))
        ax.set_ylim(0, max_value * 1.25 if max_value > 0 else 10)

        ax.grid(True, alpha=0.5, linestyle='-', axis='y', color='lightgray')
        ax.set_axisbelow(True)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.spines['bottom'].set_color('lightgray')
        ax.legend(loc='upper right', frameon=False, fontsize=10)
        plt.tight_layout()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        chart_path = f"temp/grafico_{metrica}_{''.join(series_selecionadas)}_{timestamp}.png"
        fig.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        
        return {'path': chart_path, 'title': titulo_grafico, 'description': subtitulo_grafico}
    except Exception as e:
        print(f"❌ Erro ao gerar gráfico para métrica '{metrica}': {e}")
        return {}

def create_pdf_report(charts_data: Dict[str, Any], quant_trimestre: int, polo: str, total_alunos_p1: int, total_alunos_p2: int = None) -> str:
    """
    Cria relatório PDF com página de rosto, gráficos e descrições das métricas.
    """
    try:
        print(f"🔍 DEBUG: Iniciando criação do PDF para o {quant_trimestre}º trimestre")
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        pdf_path = f"temp/relatorio_graficos_{quant_trimestre}_trimestre_{timestamp}.pdf"
        os.makedirs("temp", exist_ok=True)

        doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=50, leftMargin=50, topMargin=60, bottomMargin=60)
        
        styles = getSampleStyleSheet()
        story = []
        
        # --- Estilos ---
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=28, spaceAfter=25, alignment=1, textColor=colors.HexColor('#165b70'), fontName='Helvetica-Bold')
        subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Heading2'], fontSize=18, spaceAfter=12, alignment=1, textColor=colors.HexColor('#3d626d'), fontName='Helvetica-Bold')
        normal_style_center = ParagraphStyle('CustomNormalCenter', parent=styles['Normal'], fontSize=12, spaceAfter=10, alignment=1, textColor=colors.HexColor('#333333'), fontName='Helvetica')
        
        # NOVO ESTILO: Para a descrição da métrica abaixo do gráfico
        caption_style = ParagraphStyle('CaptionStyle', parent=styles['Normal'], fontSize=10, alignment=1, textColor=colors.HexColor('#4a4a4a'), fontName='Helvetica-Oblique', leading=14)

        # NOVO DICIONÁRIO: Com as descrições das métricas
        metric_descriptions = {
            # --- Métricas de Leitura ---
            'nl': '<b>Não Leitor (NL):</b> Refere-se ao aluno que ainda não consegue ler palavras ou sílabas, demonstrando a necessidade de intervenções focadas na alfabetização básica.',
            'ls': '<b>Leitor Silábico (LS):</b> Descreve o aluno que já compreende a escrita em segmentos (sílabas), mas ainda lê de forma pausada, juntando as sílabas para formar a palavra.',
            'lp': '<b>Leitor de Palavra (LP):</b> Indica o aluno que lê palavras inteiras com certa fluidez, mas ainda de forma segmentada, sem conectar o ritmo e a entonação de uma frase completa.',
            'lf': '<b>Leitor de Frase (LF):</b> Corresponde ao aluno que lê frases completas com fluidez e entonação adequadas, demonstrando compreensão das unidades de sentido.',
            'lsf': '<b>Leitor Silábico com Fluência (LSF):</b> Etapa intermediária onde o aluno, embora ainda decodifique silabicamente, o faz com maior rapidez, iniciando a transição para a leitura de palavras.',
            'lcf': '<b>Leitor com Fluência (LCF):</b> Representa o aluno que lê textos de forma contínua, com ritmo, entonação e precisão, compreendendo o que foi lido. É o objetivo final da alfabetização em leitura.',

            # --- Métricas de Escrita ---
            'p': '<b>Pré-Silábico (P):</b> Refere-se ao aluno que, na escrita, ainda não estabelece relação entre os sons da fala e as letras, utilizando grafismos primitivos ou letras aleatórias.',
            's': '<b>Silábico (S):</b> Descreve o aluno que já compreende que a escrita representa a fala e utiliza geralmente uma letra para representar cada sílaba.',
            's.a.': '<b>Silábico-Alfabético (S.A.):</b> Fase de transição onde o aluno alterna entre a escrita silábica e a escrita alfabética (completa) dentro da mesma palavra.',
            'a': '<b>Alfabético (A):</b> Indica o aluno que já compreendeu o sistema de escrita, sendo capaz de grafar todos os sons de uma palavra, embora ainda possa cometer erros ortográficos.',
            'o': '<b>Ortográfico (O):</b> Nível onde o aluno, além de ser alfabético, já domina as convenções ortográficas da língua (uso de RR, SS, Ç, acentuação, etc.).'
        }

        # --- PÁGINA 1: PÁGINA DE ROSTO ---
        # (O código da página de rosto continua o mesmo)
        story.append(Spacer(1, 1.5 * inch))
        story.append(Paragraph("RELATÓRIO DE LEITURA E ESCRITA", title_style))
        story.append(Spacer(1, 0.5 * inch))
        story.append(Paragraph(f"ANÁLISE COMPARATIVA DO {quant_trimestre}º TRIMESTRE", subtitle_style))
        story.append(Spacer(1, 1.5 * inch))
        story.append(Paragraph(f"<b>POLO:</b> {polo.upper()}", normal_style_center))
        story.append(Paragraph(f"<b>DATA DE EMISSÃO:</b> {datetime.now().strftime('%d/%m/%Y')}", normal_style_center))
        story.append(Spacer(1, 0.5 * inch))
        story.append(Paragraph(f"<b>TOTAL DE ALUNOS (1º Período):</b> {total_alunos_p1}", normal_style_center))
        if total_alunos_p2 is not None:
            story.append(Paragraph(f"<b>TOTAL DE ALUNOS (2º Período):</b> {total_alunos_p2}", normal_style_center))
        story.append(PageBreak())

        # --- PÁGINAS SEGUINTES: GRÁFICOS ---
        for chart_key, chart_info in charts_data.items():
            if not chart_info:
                continue

            chart_block = []
            chart_block.append(Paragraph(f"<b>{chart_info['title']}</b>", subtitle_style))
            chart_block.append(Spacer(1, 15))
            chart_block.append(Paragraph(chart_info['description'], normal_style_center))
            chart_block.append(Spacer(1, 15))

            if os.path.exists(chart_info['path']):
                img = Image(chart_info['path'], width=500, height=280)
                chart_block.append(img)
                
                # --- LÓGICA COMPLETA PARA ADICIONAR A DESCRIÇÃO DA MÉTRICA ---
                description_text = ""
                # Itera sobre todas as chaves de métricas conhecidas
                for metric_key in metric_descriptions.keys():
                    # Checa se a chave do gráfico (ex: 'nl_fund1') começa com a chave da métrica (ex: 'nl')
                    if chart_key.startswith(metric_key):
                        description_text = metric_descriptions.get(metric_key)
                        break # Para a busca assim que encontrar a correspondência

                if description_text:
                    chart_block.append(Spacer(1, 10))
                    chart_block.append(Paragraph(description_text, caption_style))
                # --- FIM DA LÓGICA ---

            else:
                chart_block.append(Paragraph(f"<i>Gráfico não disponível: {chart_info['path']}</i>", styles['Italic']))
            
            story.append(KeepTogether(chart_block))
            story.append(Spacer(1, 25))

        print(f"🔍 DEBUG: Construindo PDF...")
        doc.build(story)
        
        print(f"✅ PDF gerado com sucesso: {pdf_path}")
        return pdf_path
        
    except Exception as e:
        print(f"🔍 DEBUG: Erro na criação do PDF: {e}")
        import traceback
        traceback.print_exc()
        raise Exception(f"Erro na criação do PDF: {str(e)}")

def consolidate_data(dataframes: List[pd.DataFrame]) -> pd.DataFrame:
    """
    Consolida múltiplos DataFrames em um único
    
    Args:
        dataframes: Lista de DataFrames para consolidar
    
    Returns:
        DataFrame consolidado
    """
    try:
        if not dataframes:
            raise ValueError("Lista de DataFrames vazia")
        
        # Concatenar todos os DataFrames
        consolidated = pd.concat(dataframes, ignore_index=True)
        
        # Remover duplicatas se houver
        consolidated = consolidated.drop_duplicates()
        
        return consolidated
        
    except Exception as e:
        raise Exception(f"Erro na consolidação dos dados: {str(e)}")

def validate_excel_structure(df: pd.DataFrame) -> bool:
    """
    Valida se a estrutura do Excel está correta
    
    Args:
        df: DataFrame para validar
    
    Returns:
        True se válido, False caso contrário
    """
    required_columns = ['Nome da escola', 'Modalidade', 'Niveis de Leitura', 'Niveis de Escrita']
    
    # Verificar se todas as colunas obrigatórias existem
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        return False
    
    # Verificar se há dados
    if df.empty:
        return False
    
    # Verificar se os níveis são válidos
    niveis_validos = ['Baixo', 'Médio', 'Alto']
    
    leitura_invalidos = df[~df['Niveis de Leitura'].isin(niveis_validos)]
    escrita_invalidos = df[~df['Niveis de Escrita'].isin(niveis_validos)]
    
    if not leitura_invalidos.empty or not escrita_invalidos.empty:
        return False
    
    return True